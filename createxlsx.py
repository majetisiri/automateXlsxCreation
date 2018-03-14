import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import string
import time
start_time = time.time()

def loadXlsxFile():
	wb = openpyxl.load_workbook('specs1.xlsx')
	return wb

def getSpecsSheet(wb):
	# print wb.get_sheet_names()
	sheet=wb['specs1']
	return sheet

def deleteOutputSheetIfExists(wb):
	if 'output' in wb.sheetnames:
		std=wb['output']
		wb.remove(std)
	
def createOutputSheet(wb):
	outputSheet=wb.create_sheet(title='output')
	sheet=wb['output']
	return sheet

def copyColsFromSheets(wb,specsSheet,outputSheet):
	rows=specsSheet.max_row
	columns=specsSheet.max_column
	listab=[]
	for i in range(1, rows+1):
		listab.append([])
	for r in range(1,rows+1):
		for c in range(1,columns+1):
			e=specsSheet.cell(row=r,column=c)
			listab[r-1].append(e.value)
	for r in range(1,rows+1):
		for c in range(1,9):
			j=outputSheet.cell(row=r,column=c)
			if c ==1:
				if r==1:
					j.value="specs MCC"
				else:
					j.value=listab[r-1][3]
			elif c==2:
				if r==1:
					j.value="spec_name Website"
				else:
					j.value=listab[r-1][5]
			elif c==3:
				if r==1:
					j.value="spec_name PSN"
				else:
					func = lambda s: s[:1].lower() + s[1:] if s else ''
					new_str=listab[r-1][5].replace(" ","")
					j.value=func(new_str)
			elif c==4:
				if r==1:
					j.value="parent Name"
				else:
					j.value=listab[r-1][4]
			elif c==5:
				if r==1:
					j.value="append"
				else:
					j.value="null"
			elif c==6:
				if r==1:
					j.value="prepend"
				else:
					j.value="null"
			elif c==7:
				if r==1:
					j.value="imperial unitName"
				else:
					j.value="null"
			elif c==8:
				if r==1:
					j.value="metric unitName"
				else:
					j.value="null"
	wb.save('specs1.xlsx')

def changeBackgroundColor(wb):
	# Change background color 
	sheet=wb['output']
	for cell in sheet["1:1"]:
		cell.fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
		cell.font = Font(color="FFFFFFFF",italic=True,name="Arial",size=10)
		sheet.auto_filter.ref='A1:H1'
	wb.save('specs1.xlsx')


def getSpecName():
	wb = openpyxl.load_workbook('specs1.xlsx')
	ws= wb['output']
	rowCount=ws.max_row
	for i in range(2,rowCount+1):
		specNameCellString= "B"+str (i)
		specMCCCellString= "A"+str (i)
		specParentNameCellString= "D"+str (i)
		specName= ws[specNameCellString].value
		specMCC= ws[specMCCCellString].value
		parentName=getParentNameForSpec(specName,specMCC)
		ws[specParentNameCellString]=parentName
		wb.save('specs1.xlsx')

def getParentNameForSpec(specName,specMCC):
	filePath='specCodes/'+specMCC+'.xlsx'
	wb = openpyxl.load_workbook(filePath)
	ws= wb['Sheet1']
	rowCount = ws.max_row
	parentName=""
	for i in range(2,rowCount+1):
		specMCCCellString= "H"+str(i)
		specNameCellString= "D"+str (i)
		parentNameCellString= "E"+str (i)
		specMCCFromCodeSheet=ws[specMCCCellString].value
		specNameFromCodeSheet=ws[specNameCellString].value
		if specMCC ==specMCCFromCodeSheet:
			if specName ==specNameFromCodeSheet:
				parentName= ws[parentNameCellString].value
	return parentName

wb=loadXlsxFile()
specsSheet=getSpecsSheet(wb)
deleteOutputSheetIfExists(wb)
outputSheet=createOutputSheet(wb)
wb.save('specs1.xlsx')
copyColsFromSheets(wb,specsSheet,outputSheet)
changeBackgroundColor(wb)
getSpecName()
print("--- %s seconds ---" % (time.time() - start_time))